#!/usr/bin/env python3
# NABIS(국가균형발전종합정보시스템·산업연구원) 지역발전·혁신·창조잠재력 지수 → data/nabis.json
#   ⚠ 원자료는 '시도 단위' 시계열(발전 ~2018·혁신 ~2017·창조 ~2022). 동별로는 소속 시도값 공유(컨텍스트).
#   사용자가 지방시대위원회(NABIS) 사이트에서 받은 '시계열자료(사이트게재)' xlsx 3종을 읽는다.
import openpyxl, glob, json, os, unicodedata

DIR = os.environ.get("NABIS_DIR", os.path.join(os.path.dirname(__file__), "..", "지방시대위원회"))
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "nabis.json")


def core(s):
    s = unicodedata.normalize("NFC", str(s)).strip()
    table = {"경기도": "경기", "강원도": "강원", "강원특별자치도": "강원", "충청북도": "충북", "충청남도": "충남",
             "전라북도": "전북", "전북특별자치도": "전북", "전라남도": "전남", "경상북도": "경북", "경상남도": "경남",
             "제주특별자치도": "제주", "제주도": "제주", "세종특별자치시": "세종", "세종시": "세종"}
    if s in table:
        return table[s]
    for suf in ["특별자치도", "특별자치시", "특별시", "광역시"]:
        s = s.replace(suf, "")
    return s[:2]


def num(v):
    try:
        if v is None:
            return None
        f = float(str(v).strip())
        return round(f, 4)
    except Exception:
        return None


VALID = {"서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종", "경기", "강원", "충북", "충남", "전북", "전남", "경북", "경남", "제주"}


def latest_by_sido(path, sheet_kw):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = None
    for s in wb.sheetnames:
        if sheet_kw in unicodedata.normalize("NFC", s):
            sheet = s
            break
    if sheet is None:
        sheet = wb.sheetnames[-1]
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    # 헤더 행 = 연도가 든 첫 행(2번째 셀이 4자리 연도)
    hdr_i = 0
    for i, r in enumerate(rows[:5]):
        if r and any(str(c).strip().isdigit() and len(str(c).strip()) == 4 for c in r[1:6] if c is not None):
            hdr_i = i
            break
    header = [str(c).strip() if c is not None else "" for c in rows[hdr_i]]
    out = {}
    for r in rows[hdr_i + 1:]:
        if not r or r[0] is None:
            continue
        name = unicodedata.normalize("NFC", str(r[0]).strip())
        if name in ("전국", "계", "") or len(name) < 2:
            continue
        ck = core(name)
        if ck not in VALID:  # 권역(대경·동남·수도·충청·호남) 등 제외 — 17개 시도만
            continue
        # 가장 오른쪽(최신) '연도 컬럼'의 유효 값 (헤더가 4자리 연도인 칼럼만)
        val, yr = None, None
        for j in range(len(r) - 1, 0, -1):
            hy = header[j] if j < len(header) else ""
            if not (hy.isdigit() and len(hy) == 4):
                continue
            n = num(r[j])
            if n is not None:
                val, yr = n, hy
                break
        if val is not None:
            out[ck] = {"value": val, "year": yr}
    print(f"  [{os.path.basename(path)[:8]}] 연도컬럼 말미: {[h for h in header if h.isdigit() and len(h)==4][-4:]}")
    return out


def find(kw):
    for f in glob.glob(os.path.join(DIR, "*.xlsx")):
        nf = unicodedata.normalize("NFC", os.path.basename(f))
        if "시계열" in nf and kw in nf:
            return f
    return None


def main():
    develop = latest_by_sido(find("발전"), "총합") if find("발전") else {}
    innovate = latest_by_sido(find("혁신"), "총합") if find("혁신") else {}
    creative = latest_by_sido(find("창조"), "종합") if find("창조") else {}

    sidos = set(develop) | set(innovate) | set(creative)
    by_sido = {}
    for sd in sidos:
        d, i, c = develop.get(sd), innovate.get(sd), creative.get(sd)
        by_sido[sd] = {
            "develop": d["value"] if d else None, "devYear": d["year"] if d else None,
            "innovate": i["value"] if i else None, "innYear": i["year"] if i else None,
            "creative": c["value"] if c else None, "creYear": c["year"] if c else None,
        }
    data = {
        "source": "NABIS 국가균형발전종합정보시스템 · 산업연구원",
        "scope": "시도 단위(동별 broadcast)",
        "bySido": by_sido,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fp:
        json.dump(data, fp, ensure_ascii=False, indent=0)
    print(f"nabis.json 작성: {len(by_sido)}개 시도")
    for sd in sorted(by_sido):
        v = by_sido[sd]
        print(f"  {sd}: 발전 {v['develop']}({v['devYear']}) · 혁신 {v['innovate']}({v['innYear']}) · 창조 {v['creative']}({v['creYear']})")


if __name__ == "__main__":
    main()
